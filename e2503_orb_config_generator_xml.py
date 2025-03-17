import os
import re
from openpyxl import load_workbook
from lxml import etree

# Mapping for non-ZTS groups (from second part of resource sub‑type)
GROUP_MAP = {
    "OfflineParameter": "offlineparam",
    "CFXAPP": "cfxapp",
    "MICM": "micm",
    "FEEC": "feec",
    "SCSCF": "scscf",
    "ADMIN": "admin",
    "DeployParameter": "deploy",
    "ICSCF": "icscf",
    "FEEF": "feef",
    "CIF": "cif"
}

# Mapping for ZTS groups (when resource sub-type is like "cfx^ZTS^groupname")
ZTS_MAP = {
    "ocsp_config": ("zts_oscp", "http://www.nokia.com/zts_oscp"),
    "log-config-common": ("zts_lfs", "http://www.nokia.com/zts_lfs"),
    "log-destination-splunk": ("zts_lfs", "http://www.nokia.com/zts_lfs"),
    "log-destination-kafka": ("zts_lfs", "http://www.nokia.com/zts_lfs"),
    "log-destination-fluentd": ("zts_lfs", "http://www.nokia.com/zts_lfs"),
    "log-destination-elastic": ("zts_lfs", "http://www.nokia.com/zts_lfs"),
    "log-destination-rsyslog": ("zts_lfs", "http://www.nokia.com/zts_lfs"),
    "log-destination-sftp": ("zts_lfs", "http://www.nokia.com/zts_lfs"),
    "ss-config": ("zts_ss", "http://www.nokia.com/zts_ss"),
    "um_config": ("zts_um", "http://www.nokia.com/zts_um"),
    "cert_revocation_config": ("zts_cert", "http://www.nokia.com/zts_cert"),
    "certificate-authority": ("zts_cert", "http://www.nokia.com/zts_cert"),
    "cliserver_config": ("zts_cli", "http://www.nokia.com/zts_cli")
}

# Parameter corrections mapping.
# For example, if Excel provides "pm-nbi-prometheus-enabled", our output will use "pm-nbi-rtpm-prometheus-enabled"
PARAM_CORRECTIONS = {
    "pm-nbi-prometheus-enabled": "pm-nbi-rtpm-prometheus-enabled"
}

# Global list to capture log messages.
log_lines = []

def log(msg):
    print(msg)
    log_lines.append(msg)

def sanitize_tag(tag):
    """
    Sanitizes a string to be a valid XML tag:
      - Replaces spaces with underscores.
      - Removes characters not allowed in XML tags.
      - Ensures the tag starts with a letter or underscore.
    """
    tag = tag.replace(" ", "_")
    tag = re.sub(r"[^\w\-.]", "", tag)
    if not re.match(r"[A-Za-z_]", tag):
        tag = "_" + tag
    return tag

def normalize(name):
    """
    Normalize a string by lowercasing and removing non-alphanumeric characters.
    Used for fuzzy matching.
    """
    return "".join(ch for ch in name.lower() if ch.isalnum())

def read_excel_sheet_data(wb, sheet_name, site):
    """
    Reads data from the given sheet.
    Expected columns:
      - "NE Parameter Name"
      - "Resource Sub-Type" (e.g. "cfx^OfflineParameter" or "cfx^ZTS^ocsp_config")
      - A column whose header matches SITE (case-insensitive) holds the parameter value.
    Returns a dictionary mapping parameter names to details.
    """
    if sheet_name not in wb.sheetnames:
        log(f"Sheet '{sheet_name}' not found.")
        return {}
    ws = wb[sheet_name]
    data = {}
    headers = [cell.value.strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        param_index = headers.index("NE Parameter Name")
    except ValueError:
        log("Column 'NE Parameter Name' not found.")
        return {}
    try:
        res_subtype_index = headers.index("Resource Sub-Type")
    except ValueError:
        log("Column 'Resource Sub-Type' not found.")
        return {}
    site_index = None
    for idx, header in enumerate(headers):
        if header.lower() == site.lower():
            site_index = idx
            break
    if site_index is None:
        log(f"Site column '{site}' not found in sheet '{sheet_name}'.")
        return {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        param_name = row[param_index]
        if param_name and str(param_name).strip().lower() == "ne parameter name":
            continue
        res_subtype = row[res_subtype_index]
        value = row[site_index]
        if param_name and str(param_name).strip():
            key = str(param_name).strip()
            data[key] = {"resource_subtype": str(res_subtype).strip() if res_subtype else "", "value": value}
    return data

def aggregate_excel_data(excel_path, site):
    """
    Aggregates Excel data from sheets NP and SP (SP overrides NP).
    """
    if not os.path.exists(excel_path):
        log(f"Excel file not found at: {excel_path}")
        return {}
    wb = load_workbook(filename=excel_path, data_only=True)
    aggregated = {}
    for sheet in ["NP", "SP"]:
        sheet_data = read_excel_sheet_data(wb, sheet, site)
        aggregated.update(sheet_data)
    return aggregated

def load_xml_template(xml_template_path):
    """
    Loads the XML template and removes blank text.
    """
    if not os.path.exists(xml_template_path):
        log(f"XML template file not found at: {xml_template_path}")
        return None
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(xml_template_path, parser)
    return tree

def get_group_info(resource_subtype):
    """
    Processes the resource sub‑type.
    If it splits into exactly 2 parts (e.g. "cfx^OfflineParameter"), returns (group_name,).
    If it splits into exactly 3 parts and the second part is "ZTS", returns ("ZTS", group_name).
    Otherwise, returns (parts[1],) (ignoring extra parts).
    """
    parts = resource_subtype.split("^")
    if len(parts) < 2:
        return None
    if len(parts) == 3 and parts[1].strip().upper() == "ZTS":
        return ("ZTS", parts[2].strip())
    else:
        return (parts[1].strip(),)

def clean_xml(element):
    """
    Recursively removes child elements that have text exactly 'n/a' (case-insensitive)
    and that have no child elements.
    """
    for child in list(element):
        clean_xml(child)
        if child.text is not None and child.text.strip().lower() == "n/a" and len(child) == 0:
            element.remove(child)

def update_xml_template(tree, excel_data):
    """
    Updates the XML template based on Excel parameters.
    • For non‑ZTS groups, container is <cfx:cfx>.
    • For ZTS groups, container is <zts_cm:ZTS>.
    • Each Excel parameter is placed in the group determined by its Resource Sub‑Type.
    • New element tag names are sanitized and corrected using PARAM_CORRECTIONS.
    """
    root = tree.getroot()
    nsmap = root.nsmap

    # Container for non-ZTS groups: <cfx:cfx>
    container = tree.xpath("//cfx:cfx", namespaces=nsmap)
    if not container:
        log("Container element <cfx:cfx> not found in template.")
        return tree
    container = container[0]

    # Container for ZTS groups: <zts_cm:ZTS>
    zts_container = tree.xpath("//zts_cm:ZTS", namespaces=nsmap)
    if zts_container:
        zts_container = zts_container[0]
    else:
        zts_container = container

    # Group Excel data by processed group info.
    groups = {}
    for param, details in excel_data.items():
        res_type = details.get("resource_subtype", "")
        grp_info = get_group_info(res_type)
        if not grp_info:
            log(f"Skipping parameter '{param}' (invalid Resource Sub-Type: '{res_type}').")
            continue
        groups.setdefault(grp_info, {})[param] = details.get("value")

    # Process each group.
    for grp_key, params in groups.items():
        if len(grp_key) == 1:
            # Non-ZTS group.
            group_name = grp_key[0]
            if group_name in GROUP_MAP:
                desired_prefix = GROUP_MAP[group_name]
                desired_ns = nsmap.get(desired_prefix, "")
            else:
                desired_prefix = group_name.lower()
                desired_ns = nsmap.get(desired_prefix, "")
            expected_local = group_name
            group_container = container
        elif len(grp_key) == 2:
            # ZTS group.
            group_name = grp_key[1]
            if group_name in ZTS_MAP:
                desired_prefix, desired_ns = ZTS_MAP[group_name]
            else:
                desired_prefix = group_name.lower()
                desired_ns = nsmap.get(desired_prefix, "")
            expected_local = group_name
            group_container = zts_container
        else:
            group_name = grp_key[0]
            desired_prefix = GROUP_MAP.get(group_name, group_name.lower())
            desired_ns = nsmap.get(desired_prefix, "")
            expected_local = group_name
            group_container = container

        xpath_expr = f"./*[local-name()='{expected_local}'"
        if desired_ns:
            xpath_expr += f" and namespace-uri()='{desired_ns}'"
        xpath_expr += "]"
        parent_elements = group_container.xpath(xpath_expr, namespaces=nsmap)
        if not parent_elements:
            if desired_ns:
                group_tag = f"{{{desired_ns}}}{expected_local}"
            else:
                group_tag = expected_local
            parent_element = etree.SubElement(group_container, group_tag)
            log(f"Created new group element '{expected_local}' with prefix '{desired_prefix}'.")
        else:
            parent_element = parent_elements[0]

        # Build mapping of normalized child names.
        child_map = {}
        for child in parent_element:
            child_name = etree.QName(child).localname
            child_map[normalize(child_name)] = child

        for param, new_value in params.items():
            norm_param = normalize(param)
            if norm_param in child_map:
                child_el = child_map[norm_param]
                child_el.text = str(new_value) if new_value is not None else ""
            else:
                safe_param = sanitize_tag(param)
                corrected_param = PARAM_CORRECTIONS.get(safe_param, safe_param)
                if desired_ns:
                    new_tag = f"{{{desired_ns}}}{corrected_param}"
                else:
                    new_tag = corrected_param
                new_child = etree.SubElement(parent_element, new_tag)
                new_child.text = str(new_value) if new_value is not None else ""
                log(f"Created new parameter element '{param}' (tag '{corrected_param}') under group '{expected_local}'.")
    return tree

def update_deploy_group(tree):
    """
    Checks the deploy group (<deploy:DeployParameter>) under <cfx:cfx> and ensures required parameters.
    Required defaults:
      - <deploy:ZtsLcmUsername> is set to "zts1user"
      - <deploy:ZtsLcmPassword> is set to "********"
      - <deploy:CncsUMAdminPassword> is set to "none"
    If <deploy:TrafficFileName> is missing, prompts the user for its value.
    Returns a dictionary of deploy parameters.
    """
    root = tree.getroot()
    nsmap = root.nsmap
    container = tree.xpath("//cfx:cfx", namespaces=nsmap)
    if not container:
        log("Container <cfx:cfx> not found for deploy group update.")
        return {}
    container = container[0]
    xpath_expr = "./*[local-name()='DeployParameter']"
    deploy_group = container.xpath(xpath_expr, namespaces=nsmap)
    if not deploy_group:
        desired_ns = nsmap.get("deploy", "")
        tag = f"{{{desired_ns}}}DeployParameter" if desired_ns else "DeployParameter"
        deploy_group = [etree.SubElement(container, tag)]
        log("Created new deploy group element 'DeployParameter'.")
    deploy_group = deploy_group[0]
    required = {
        "ZtsLcmUsername": "zts1user",
        "ZtsLcmPassword": "********",
        "CncsUMAdminPassword": "none"
    }
    existing = {etree.QName(child).localname for child in deploy_group}
    for param, default_val in required.items():
        if param not in existing:
            desired_ns = nsmap.get("deploy", "")
            tag = f"{{{desired_ns}}}{param}" if desired_ns else param
            new_child = etree.SubElement(deploy_group, tag)
            new_child.text = default_val
            log(f"Added deploy parameter '{param}' with default value '{default_val}'.")
    if "TrafficFileName" not in existing:
        user_val = input("Enter value for TrafficFileName: ").strip()
        desired_ns = nsmap.get("deploy", "")
        tag = f"{{{desired_ns}}}TrafficFileName" if desired_ns else "TrafficFileName"
        new_child = etree.SubElement(deploy_group, tag)
        new_child.text = user_val
        log(f"Added deploy parameter 'TrafficFileName' with value '{user_val}'.")
    deploy_vars = {}
    for child in deploy_group:
        deploy_vars[etree.QName(child).localname] = child.text
    return deploy_vars

def generate_config_for_site(excel_path, xml_template_path, site, output_dir):
    """
    Processes the SITE:
      - Aggregates Excel data.
      - Updates the XML template.
      - Updates the deploy group.
      - Cleans elements with text "n/a".
      - Writes the output XML, log, and var files.
    Output file names default to:
       <SITE>_config.xml, <SITE>_config.log, and <SITE>_config.vars in the output directory.
    """
    global log_lines
    log_lines = []  # Reset log for this site
    log(f"Processing SITE: {site}")
    excel_data = aggregate_excel_data(excel_path, site)
    if not excel_data:
        log("No valid Excel data found. Skipping this SITE.")
        return
    tree = load_xml_template(xml_template_path)
    if tree is None:
        return
    updated_tree = update_xml_template(tree, excel_data)
    deploy_vars = update_deploy_group(updated_tree)
    clean_xml(updated_tree.getroot())
    xml_filename = os.path.join(output_dir, f"{site}_config.xml")
    log_filename = os.path.join(output_dir, f"{site}_config.log")
    var_filename = os.path.join(output_dir, f"{site}_config.vars")
    updated_tree.write(xml_filename, pretty_print=True, xml_declaration=True, encoding="UTF-8")
    log(f"Generated XML configuration saved to: {xml_filename}")
    with open(log_filename, "w", encoding="UTF-8") as lf:
        lf.write("\n".join(log_lines))
    log(f"Log file saved to: {log_filename}")
    with open(var_filename, "w", encoding="UTF-8") as vf:
        for key, value in deploy_vars.items():
            vf.write(f"{key}={value}\n")
    log(f"Var file saved to: {var_filename}")

def generate_config(excel_path, xml_template_path, sites, output_dir):
    """
    Processes each SITE in the provided list.
    """
    for site in sites:
        generate_config_for_site(excel_path, xml_template_path, site.strip(), output_dir)

def main():
    excel_path = input("Enter the path to the Excel file (with sheets NP and SP): ").strip()
    xml_template_path = input("Enter the path to the XML Template file: ").strip()
    sites_input = input("Enter comma-separated SITE column names to match in the Excel file: ").strip()
    output_dir = input("Enter the output directory path (where configuration, log, and var files will be saved): ").strip()
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    sites = [s.strip() for s in sites_input.split(",") if s.strip()]
    if not sites:
        print("No SITE names provided. Exiting.")
        return
    for f in [excel_path, xml_template_path]:
        if not os.path.exists(f):
            print(f"File not found: {f}")
            return
    generate_config(excel_path, xml_template_path, sites, output_dir)

if __name__ == "__main__":
    main()